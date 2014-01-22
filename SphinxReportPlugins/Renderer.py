import os, sys, re, shelve, traceback, pickle, types, itertools, collections

# so that arange is available in eval
import numpy
import numpy.ma
from numpy import *
from math import *

from SphinxReport.ResultBlock import ResultBlock, EmptyResultBlock, ResultBlocks
from collections import OrderedDict as odict
from SphinxReport.DataTree import path2str, tree2table
from SphinxReport.Component import *
from SphinxReport import Utils, DataTree
from SphinxReport import CorrespondenceAnalysis

import StringIO

from docutils.parsers.rst import directives

import pandas
import pandas.core.reshape

# for output of work books
# import xlwt
import openpyxl
from openpyxl.cell import get_column_letter

class Renderer(Component):
    """Base class of renderers that render data into restructured text.

    The subclasses define how to render the data by overloading the
    :meth:`render` method.

    When called, a Renderer and its subclasses will return blocks of
    restructured text. Images are automatically collected from matplotlib
    and inserted at place-holders.

    This class adds the following options to the :term:`report` directive.

       :term:`groupby`: group data by :term:`track` or :term:`slice`.
    """
    # plugin fields
    capabilities = ["render"]

    options = ( ('format', directives.unchanged), 
                ('split-at', directives.nonnegative_int),
                ('split-always', directives.unchanged), )

    # required levels in DataTree
    nlevels = None

    # default number format
    format = "%i"

    # split number of tracks
    split_at = 0

    # tracks always to include if plot is split into
    # several plots.
    split_always = []

    def __init__(self, *args, **kwargs ):
        """create an Renderer object using an instance of 
        a :class:`Tracker.Tracker`.
        """

        self.debug("%s: starting renderer '%s'"% (id(self), str(self)))

        try: self.format = kwargs["format"]
        except KeyError: pass

        self.split_at = int(kwargs.get( "split-at", 0))

        if "split-always" in kwargs:
            self.split_always = kwargs["split-always"].split(',')
        else:
            self.split_always = None

    def __call__(self, dataseries, path ):
        '''iterate over leaves/branches in data structure.

        This method will call the :meth:`render` method for 
        each leaf/branch at level :attr:`nlevels`.
        '''
        if self.nlevels == None: raise NotImplementedError("incomplete implementation of %s" % str(self))

        try:
            labels = dataseries.index.levels
            paths = dataseries.index.unique()
        except AttributeError:
            labels = ['dummy1'] 
            paths = ['dummy1']

        result = ResultBlocks()

        if len(labels) < self.nlevels:
            self.warn( "at %s: expected at least %i levels - got %i: %s" %\
                           (str(path), self.nlevels, len(labels), str(labels)) )
            result.append( EmptyResultBlock( title = path2str(path) ) )
            return result

        result.extend( self.render( dataseries, path ) )

        # for p in paths:
        #     work = DataTree.getLeaf( data, p )
        #     if not work: continue
        #     if self.split_at:
        #         k = list(work.keys())
        #         # select tracks to always add to split 
        #         if self.split_always:
        #             always = [ x for x, y in itertools.product( k, self.split_always) if re.search( y, x ) ]
                    
        #         for z, x in enumerate(range( 0, len(k), self.split_at)) :
        #             if self.split_always:
        #                 w = odict( [ (xx, work[xx]) for xx in always ] )
        #             else:
        #                 w = odict()

        #             w.update( odict( [ (xx, work[xx]) for xx in k[x:x+self.split_at] ] ) )
        #             try:
        #                 result.extend( self.render( w, path + p + (str(z),) ) )
        #             except:
        #                 self.warn("exeception raised in rendering for path: %s" % str(path+p+str((z,))))
        #                 raise 
        #     else:
        #         try:
        #             result.extend( self.render( work, path + p ) )
        #         except:
        #             self.warn("exeception raised in rendering for path: %s" % str(path+p))
        #             raise 
            
        return result

    def toString( self, value ):
        '''returns a number as string
        
        If not a number, return empty string.'''
        
        try: return self.format % value
        except TypeError: return ""

class TableBase( Renderer ):
    '''base classes for tables and matrices.'''

    options = Renderer.options +\
        ( ('force', directives.unchanged), 
          ('separate', directives.unchanged), 
          ('max-rows', directives.length_or_unitless),
          ('max-cols', directives.length_or_unitless),
          )
    
    max_rows = 50
    max_cols = 20

    def __init__( self, *args, **kwargs ):
        Renderer.__init__(self, *args, **kwargs )

        self.force = "force" in kwargs            
        self.separate = "separate" in kwargs
        self.max_rows = kwargs.get( "max-rows", 50 )
        self.max_cols = kwargs.get( "max-cols", 20 )

    def asFile( self, dataframe, row_headers, col_headers, title ):
        '''save the table as HTML file.

        Multiple files of the same Renderer/Tracker combination are distinguished 
        by the title.
        '''

        self.debug("%s: saving %i x %i table as file'"% (id(self), 
                                                         len(row_headers), 
                                                         len(col_headers)))
        lines = []
        lines.append("`%i x %i table <#$html %s$#>`__" %\
                     (len(row_headers), len(col_headers),
                      title) )

        out = StringIO.StringIO()
        dataframe.to_csv( out )
        lines = out.getvalue().split("\n")            

        r = ResultBlock( "\n".join(lines) + "\n", title = title)

        # create an html table
        data = ["<table>"]
        data.append( "<tr><th></th><th>%s</th></tr>" % "</th><th>".join( map(str,lines[0])) )
        data.extend( ["<tr><td>%s</td></tr>" % \
                          ("</td><td>".join(x.split(","))) for x in lines[1:]] )
        data.append( "</table>\n" )

        # substitute links
        data = [ re.sub("`(.*?(?:\".+\"|\'.+\')?.*?)\s<(.*?(?:\".+\"|\'.+\')?.*?)>`_", r'<a href="\2">\1</a>', x) \
                     for x in data ]

        r.html = "\n".join( data )

        return r

    def asSpreadSheet( self, matrix, row_headers, col_headers, title ):
        '''save the table as an xls file.

        Multiple files of the same Renderer/Tracker combination are distinguished 
        by the title.
        '''
        
        self.debug("%s: saving %i x %i table as spread-sheet'"% (id(self), 
                                                                 len(row_headers), 
                                                                 len(col_headers)))
        lines = []
        lines.append("`%i x %i table <#$xls %s$#>`__" %\
                     (len(row_headers), len(col_headers),
                      title) )
        lines.append( "" )
        
        r = ResultBlock( "\n".join(lines), title = title)

        # create an html table
        wb = openpyxl.Workbook( optimized_write = True)

        ws = wb.create_sheet()
        # patch: maximum title length seems to be 31
        ws.title = title[:31]

        ws.append( [""] + list(col_headers) )
        for x,row in enumerate( matrix ):
            ws.append( [row_headers[x]] + row )

        r.xls = wb

        self.debug("%s: saved %i x %i table as spread-sheet'"% (id(self), 
                                                                len(row_headers), 
                                                                len(col_headers)))

        return r

class Table( TableBase ):
    '''a basic table. 

    Values are either text or converted to text.

    If the has more rows than :attr:`max_rows` or more columns
    than :attr:`max_cols`, a placeholder will be inserted pointing
    towards a file.

    The attribute :attr:`large` determines where large tables are written
    to. The default is html. Alternative values are ``xls`` for excel spread-sheets.

    '''
    options = TableBase.options +\
        ( ('transpose', directives.unchanged),
          ('add-rowindex', directives.unchanged),
          ('add-percent', directives.unchanged),
          ('head', directives.length_or_unitless),
          ('large', directives.unchanged),
          ('preview', directives.unchanged) )

    nlevels = -1

    transpose = False

    def __init__( self, *args, **kwargs ):
        TableBase.__init__(self, *args, **kwargs )

        self.transpose = "transpose" in kwargs
        self.add_rowindex = "add-rowindex" in kwargs
        self.large = kwargs.get( "large", "html")
        self.preview = "preview" in kwargs
        self.add_percent = kwargs.get( 'add-percent', None )
        self.head = int( kwargs.get( 'head', 0 ) )

    def modifyTable( self, matrix, row_headers, col_headers ):
        '''modify table if required, for example adding percentages.'''
        
        if self.add_percent:
            parts = self.add_percent.split(";")
            for part in parts:
                total, other_col = None, None
                if "," in part: 
                    column, method = part.split(",")
                    if method in col_headers:
                        other_col = col_headers.index( method )
                    else:
                        try: total = float( method )
                        except ValueError:
                            raise ValueError("can't compute total from expression `%s` - missing column?" % method )
                else:
                    column = part
                
                if column not in col_headers:
                    raise ValueError("unknown column `%s` to add-percent" % (column))

                col = col_headers.index( column )
                values = [ float(x[col]) for x in matrix ]

                if other_col:
                    for row in matrix:
                        row.insert( col+1, 100.0 * float(row[col]) / float(row[other_col]) )
                else:
                    if total == None: total = sum(values)
                    for row in matrix:
                        row.insert( col+1, 100.0 * float(row[col]) / total )

                col_headers.insert( col+1, "%s / %%" % column )

        return matrix, row_headers, col_headers

    def __call__(self, dataseries, path):
        
        # modify table (adding/removing columns) according to user options
        # matrix, row_headers, col_headers = self.modifyTable( matrix, row_headers, col_headers )

        title = path2str(path)

        results = ResultBlocks()
        try:
            dataframe = dataseries.unstack()
        except pandas.core.reshape.ReshapeError:
            dataframe = pandas.DataFrame(dataseries)

        row_headers = dataframe.index
        col_headers = dataframe.columns

        # do not output large matrices as rst files
        if self.separate or (not self.force and 
                             (len(row_headers) > self.max_rows or 
                              len(col_headers) > self.max_cols)):
            if self.large == "xls":
                results.append( self.asSpreadSheet( dataframe, row_headers, col_headers, title ) )
            else:
                results.append( self.asFile( dataframe, row_headers, col_headers, title ) )

            if self.preview:
                row_headers = row_headers[:self.max_rows]
                col_headers = col_headers[:self.max_cols]
                matrix = [ x[:self.max_cols] for x in matrix[:self.max_rows] ]
            else:
                return results

        out = StringIO.StringIO()
        dataframe.to_csv( out )
        lines = []
        lines.append( ".. csv-table:: %s" % title )
        lines.append( "   :class: sortable" )
        
        if self.add_rowindex:
            lines.append( '   :header: "row", "", "%s" ' % '","'.join( map(str, col_headers) ) )
            lines.append( '' )

            x = 0
            for header, line in zip( row_headers, matrix ):
                x += 1
                lines.append( '   %i,"%s","%s"' % (x, str(header), '","'.join( map(str, line) ) ) )

        else:
            l = out.getvalue().split("\n")            
            lines.append( '   :header: %s' % l[0] )
            lines.append( '' )
            lines.extend( ['   %s' % x for x in l[1:] ] )

        lines.append( "") 
        
        results.append( ResultBlock( "\n".join(lines), title = title) )

        return results

class RstTable( Table ):
    '''an rst formatted table.'''

    def __call__(self, data, path):
        
        matrix, row_headers, col_headers = self.buildTable( data )

        title = path2str(path)

        if matrix == None: 
            return ResultBlocks( ResultBlock( "\n".join(lines), title = title) )

        # do not output large matrices as rst files
        if self.separate or (not self.force and 
                             (len(row_headers) > self.max_rows or len(col_headers) > self.max_cols)):
            return ResultBlocks( self.asFile( matrix, row_headers, col_headers, title ),
                                 title = title )

        lines = []
        
        # add row and column headers
        matrix.insert( 0, col_headers )
        max_widths = [ max( len(x) for x in row_headers ) ]
        max_widths.extend( [ max([len(str(row[x])) for row in matrix]) for x in range(len(col_headers)) ] )

        separator = "+" + "+".join( [ "-" * (x + 2) for x in max_widths ] ) + "+"
        format = "|" + "|".join( [" %%%is " % x for x in max_widths] ) + "|"

        lines.append( separator )
        lines.append( format % tuple( [""] + list(map(str,col_headers)) ))
        lines.append( separator )

        for h, row in zip(row_headers,matrix[1:]):
            lines.append( format % tuple(map(str,[h] + row) ))
            lines.append( separator )

        return ResultBlocks( ResultBlock( "\n".join(lines), title = title) )

class Glossary( Table ):
    """output a table in the form of a glossary."""

    def __call__(self, data, path ):

        lines = []
        matrix, row_headers, col_headers = self.buildTable( data )

        if matrix == None: return lines

        title = path2str(path)

        lines.append( ".. glossary::" )

        lines.append( "") 

        for header, line in zip( row_headers, matrix ):
            txt = "\n".join( line )
            txt = "\n      ".join( [ x.strip() for x in txt.split("\n" ) ] )
            lines.append( '   %s\n      %s' % ( header, txt ) )

        lines.append( "") 

        return ResultBlocks( "\n".join(lines), title = title )

class MatrixBase:
    '''base class for matrices.

    This base class provides utility functions for rectangular 2D matrices.

    It implements column-wise and row-wise transformations.

    This class adds the following options to the :term:`report` directive.

        :term:`transform-matrix`: apply a matrix transform. Possible choices are:

           * *correspondence-analysis*: use correspondence analysis to permute rows/columns 
           * *normalized-col-total*: normalize by column total
           * *normalized-col-max*: normalize by column maximum,
           * *normalized-col-first*: normalize by first column. The first column is then removed.
           * *normalized-row-total*: normalize by row total
           * *normalized-row-max*: normalize by row maximum
           * *normalized-row-first*: normalize by first row. The first row is then removed.
           * *normalized-total*: normalize over whole matrix
           * *normalized-max*: normalize over whole matrix
           * *sort* : sort matrix rows and columns alphanumerically.
           * filter-by-rows: only take columns that are also present in rows
           * filter-by-cols: only take columns that are also present in cols
           * square: make square matrix (only take rows and columns present in both)
           * *add-row-total* : add the row total at the bottom
           * *add-column-total* : add the column total as a last column

    Requires two levels:

       rows[dict] / columns[dict] / value

    All values need to be numerical.
    '''

    def __init__( self, *args, **kwargs ):
    
        self.mMapKeywordToTransform = {
            "correspondence-analysis": self.transformCorrespondenceAnalysis,
            "transpose": self.transformTranspose,
            "normalized-row-total" : self.transformNormalizeRowTotal,
            "normalized-row-max" : self.transformNormalizeRowMax,
            "normalized-row-first" : self.transformNormalizeRowFirst,
            "normalized-col-total" : self.transformNormalizeColumnTotal,
            "normalized-col-max" : self.transformNormalizeColumnMax,
            "normalized-col-first" : self.transformNormalizeColumnFirst ,
            "normalized-total" : self.transformNormalizeTotal,
            "normalized-max" : self.transformNormalizeMax,
            "symmetric-max" : self.transformSymmetricMax,
            "symmetric-min" : self.transformSymmetricMin,
            "symmetric-avg" : self.transformSymmetricAverage,
            "symmetric-sum" : self.transformSymmetricSum,
            "filter-by-rows" : self.transformFilterByRows,
            "filter-by-cols" : self.transformFilterByColumns,
            "square" : self.transformSquare,
            "add-row-total" : self.transformAddRowTotal,
            "add-column-total" : self.transformAddColumnTotal,
            "sort" : self.transformSort,
        }
        
        self.converters = []        
        if "transform-matrix" in kwargs:
            for kw in [x.strip() for x in kwargs["transform-matrix"].split(",")]:
                if kw.startswith( "normalized" ): self.format = "%6.4f"
                try:
                    self.converters.append( self.mMapKeywordToTransform[ kw ] )
                except KeyError:
                    raise ValueError("unknown matrix transformation %s, possible values are: %s" \
                                         % (kw, ",".join( sorted(self.mMapKeywordToTransform.keys()) ) ) )

    def transformAddRowTotal( self, matrix, row_headers, col_headers ):
        '''add row total to the matrix.'''
        nrows, ncols = matrix.shape

        totals = numpy.zeros( nrows )
        for x in range(nrows): totals[x] = sum(matrix[x,:])
        col_headers.append( "total" )
        totals.shape=(nrows,1)
        return numpy.hstack( numpy.hsplit(matrix,ncols) + [totals,] ), row_headers, col_headers

    def transformAddColumnTotal( self, matrix, row_headers, col_headers ):
        raise NotImplementedError

    def transformFilterByRows( self, matrix, row_headers, col_headers ):
        """only take columns that are also present in rows"""
        take = [ x for x in range( len(col_headers) ) if col_headers[x] in row_headers ]
        return matrix.take( take, axis=1), row_headers, [col_headers[x] for x in take ]

    def transformFilterByColumns( self, matrix, row_headers, col_headers ):
        """only take rows that are also present in columns"""
        take = [ x for x in range( len(row_headers) ) if row_headers[x] in col_headers ]
        return matrix.take( take, axis=0), [row_headers[x] for x in take ], col_headers 

    def transformSquare( self, matrix, row_headers, col_headers ):
        """only take rows and columns that are present in both giving a square matrix."""
        take = set(row_headers).intersection( set(col_headers) )
        row_indices = [ x for x in range( len(row_headers) ) if row_headers[x] in take ]
        col_indices = [ x for x in range( len(col_headers) ) if col_headers[x] in take ]
        m1 = matrix.take( row_indices, axis=0)
        m2 = m1.take( col_indices, axis=1)

        return m2, [row_headers[x] for x in row_indices], [col_headers[x] for x in col_indices]

    def transformTranspose( self, matrix, row_headers, col_headers ):
        """transpose the matrix."""
        return numpy.transpose( matrix ), col_headers, row_headers

    def transformCorrespondenceAnalysis( self, matrix, row_headers, col_headers ):
        """apply correspondence analysis to a matrix.
        """

        if len(row_headers) <= 1 or len(col_headers) <= 1:
            self.warn( "correspondence analysis skipped for matrices with single row/column" )
            return matrix, row_headers, col_headers

        try:
            row_indices, col_indices =  CorrespondenceAnalysis.GetIndices( matrix )
        except ValueError as msg:
            return matrix, row_headers, col_headers            

        map_row_new2old = numpy.argsort(row_indices)
        map_col_new2old = numpy.argsort(col_indices)
        
        matrix, row_headers, col_headers =  CorrespondenceAnalysis.GetPermutatedMatrix( matrix,
                                                                                        map_row_new2old,
                                                                                        map_col_new2old,
                                                                                        row_headers = row_headers,
                                                                                        col_headers = col_headers)



        return matrix, row_headers, col_headers

    def transformSort( self, matrix, row_headers, col_headers ):
        """apply correspondence analysis to a matrix.
        """

        map_row_new2old = [x[0] for x in sorted(enumerate( row_headers ), key=lambda x: x[1])]
        map_col_new2old = [x[0] for x in sorted(enumerate( col_headers ), key=lambda x: x[1])]

        matrix, row_headers, col_headers =  CorrespondenceAnalysis.GetPermutatedMatrix( matrix,
                                                                                        map_row_new2old,
                                                                                        map_col_new2old,
                                                                                        row_headers = row_headers,
                                                                                        col_headers = col_headers)


        return matrix, row_headers, col_headers

    def transformSymmetricMax( self, matrix, rows, cols ):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError( "matrix not square - can not be symmetrized" )
        
        for x in range(len(rows)):
            for y in range(x+1,len(cols)):
                matrix[x,y] = matrix[y,x] = max(matrix[x,y], matrix[y,x])

        return matrix, rows, cols

    def transformSymmetricMin( self, matrix, rows, cols ):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError( "matrix not square - can not be symmetrized" )
        
        for x in range(len(rows)):
            for y in range(x+1,len(cols)):
                matrix[x,y] = matrix[y,x] = min(matrix[x,y], matrix[y,x])

        return matrix, rows, cols

    def transformSymmetricSum( self, matrix, rows, cols ):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError( "matrix not square - can not be symmetrized" )
        
        for x in range(len(rows)):
            for y in range(x+1,len(cols)):
                matrix[x,y] = matrix[y,x] = sum(matrix[x,y], matrix[y,x])

        return matrix, rows, cols

    def transformSymmetricAverage( self, matrix, rows, cols ):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError( "matrix not square - can not be symmetrized" )
        
        for x in range(len(rows)):
            for y in range(x+1,len(cols)):
                matrix[x,y] = matrix[y,x] = sum(matrix[x,y], matrix[y,x]) / 2

        return matrix, rows, cols

    def transformNormalizeTotal( self, matrix, rows, cols ):
        """normalize a matrix by the total.

        Returns the normalized matrix.
        """
        t = matrix.sum()
        matrix /= t
        return matrix, rows, cols

    def transformNormalizeMax( self, matrix, rows, cols ):
        """normalize a matrix by the max.

        Returns the normalized matrix.
        """
        t = matrix.max()
        matrix /= t
        return matrix, rows, cols

    def transformNormalizeRowTotal( self, matrix, rows, cols ):
        """normalize a matrix row by the row total.

        Returns the normalized matrix.
        """
        nrows, ncols = matrix.shape

        for x in range(nrows) :
            m = sum(matrix[x,:])
            if m != 0:
                for y in range(ncols):
                    matrix[x,y] /= m
        return matrix, rows, cols

    def transformNormalizeRowMax( self, matrix, rows, cols ):
        """normalize a matrix row by the row maximum.

        Returns the normalized matrix.
        """
        nrows, ncols = matrix.shape

        for x in range(nrows) :
            m = max(matrix[x,:])
            if m != 0:
                for y in range(ncols):
                    matrix[x,y] /= m
        return matrix, rows, cols

    def transformNormalizeRowFirst( self, matrix, rows, cols ):
        """normalize a matrix row by the row maximum.

        Returns the normalized matrix.
        """
        nrows, ncols = matrix.shape

        for x in range(1, nrows) :
            for y in range(ncols):
                m = matrix[0,y]
                if m != 0: matrix[x,y] /= m
        matrix = numpy.delete( matrix, 0, 0 )
        return matrix, rows[1:], cols

    def transformNormalizeColumnTotal( self, matrix, rows, cols ):
        """normalize a matrix by the column total.

        Returns the normalized matrix."""
        nrows, ncols = matrix.shape
        totals = [sum( matrix[:,y] ) for y in range(ncols) ]
        for x in range(nrows):
            for y in range(ncols):
                m = totals[y]
                if m != 0: matrix[x,y] /= m
        return matrix, rows, cols

    def transformNormalizeColumnFirst( self, matrix, rows, cols ):
        """normalize a matrix by the first column.

        Removes the first column.

        Returns the normalized matrix."""
        nrows, ncols = matrix.shape
        for x in range(nrows):
            m = matrix[x,0]
            for y in range(1,ncols):
                if m != 0: matrix[x,y] /= m
        matrix = numpy.delete( matrix, 0, 1)
        return matrix, rows, cols[1:]

    def transformNormalizeColumnMax( self, matrix, rows, cols ):
        """normalize a matrix by the column maximum

        Returns the normalized matrix.
        """
        nrows, ncols = matrix.shape

        for y in range(ncols) :
            m = max(matrix[:,y])
            if m != 0:
                for x in range(nrows):
                    matrix[x,y] /= m
        return matrix, rows, cols

    def render( self, work, path ):
        """render the data.
        """

        results = ResultBlocks( title = path )

        matrix, rows, columns = self.buildMatrix( work )
        
        title = path2str(path)

        if len(rows) == 0:
            return ResultBlocks( ResultBlock( "", title = title) )

        # do not output large matrices as rst files
        # separate and force need to be mixed in.
        if self.separate or (not self.force and (len(rows) > self.max_rows or len(columns) > self.max_cols)):
            return ResultBlocks( self.asFile( [ [ self.toString(x) for x in r ] for r in matrix ], 
                                              rows, 
                                              columns, 
                                              title ),
                                 title = path )

        lines = []
        lines.append( ".. csv-table:: %s" % title )
        lines.append( "   :class: sortable" )
        lines.append( '   :header: "track","%s" ' % '","'.join( columns ) )
        lines.append( '')
        for x in range(len(rows)):
            lines.append( '   "%s","%s"' % (rows[x], '","'.join( [ self.toString(x) for x in matrix[x,:] ] ) ) )
        lines.append( "") 
        if not path: subtitle = ""
        else: subtitle = path2str(path)

        results.append( ResultBlock( "\n".join(lines), title = subtitle ) )

        return results

class TableMatrix(TableBase, MatrixBase):    
    """A table with numerical columns.

       rows[dict] / columns[dict] / value

    All values need to be numerical.
    """

    nlevels = 2

    options = TableBase.options +\
        ( ('transform-matrix', directives.unchanged), )

    def __init__(self, *args, **kwargs):

        TableBase.__init__(self, *args, **kwargs )
        MatrixBase.__init__(self, *args, **kwargs )

    def buildMatrix( self, 
                     dataseries, 
                     missing_value = 0, 
                     apply_transformations = True,
                     take = None,
                     ignore = None,
                     dtype = numpy.float ):
        """build a matrix from work, a two-level nested dictionary.

        If *take* is given, then the matrix will be built from
        level 3, taking *take* from the deepest level only.
        
        If *ignore* is given, columns in ignore will be ignore. If the
        matrix is built from level 3 and no *take* is specified, ignore 
        will also applied to level 3 and the first remaining field is taken.

        This method will also apply conversions if apply_transformations
        is set.
        """

        dataframe = dataseries.unstack()
        rows = list(dataframe.index)
        columns = list(dataframe.columns)

        matrix = dataframe.as_matrix() 

        if self.converters and apply_transformations:
            for converter in self.converters: 
                self.debug("applying converter %s" % converter)
                matrix, rows, columns = converter(matrix, rows, columns)

        # convert rows/columns to str (might be None)
        rows = [ str(x) for x in rows ]
        columns = [ str(x) for x in columns ]
        
        return matrix, rows, columns

        # labels = DataTree.getPaths( work )
        # levels = len(labels)

        # rows, columns = labels[:2]

        # if ignore != None: 
        #     columns = [x for x in columns if x not in ignore ]

        # # if take is specified, this takes priority
        # if take:
        #     if levels == 3: 
        #         if take not in labels[-1]: raise ValueError( "no data on `%s`" % take )
        #         take_f = lambda row,column: work[row][column][take]
        #     elif levels == 2:
        #         take_f = lambda row,column: work[row][take]
        #         columns = [take]
        #     else:
        #         raise ValueError( "expected two or three levels, got %i: '%s'" % (levels, labels) )
        # else:
        #     if levels == 3:
        #         take = [ x for x in labels[-1] if x not in ignore ]
        #         if len(take) != 1:
        #             self.warn( "received data with three level, but third level is ambiguous, taking first of: %s" % take)
        #         take=take[0]
        #         take_f = lambda row, column: work[row][column][take]
        #     elif levels == 2: 
        #         take_f = lambda row,column: work[row][column]
        #     else:
        #         raise ValueError( "expected two levels, got %i: %s" % (levels, str(labels) ))

        # self.debug("creating matrix: taking=%s, ignoring=%s" % (take,ignore))
        # matrix = numpy.array( [missing_value] * (len(rows) * len(columns) ), dtype )
        # matrix.shape = (len(rows), len(columns) )
        # self.debug("constructing matrix")
        # for x,row in enumerate(rows):
        #     for y, column in enumerate(columns):
        #         # missing values from DataTree
        #         try:
        #             v = take_f( row, column )
        #         except KeyError:
        #             continue

        #         # empty values from DataTree
        #         try:
        #             if len(v) == 0: continue
        #         except TypeError:
        #             pass

        #         # convert
        #         try:
        #             matrix[x,y] = v
        #         except ValueError as msg:
        #             raise ValueError( "malformatted data: expected scalar, got '%s'; msg=%s" % (str(work[row][column]),msg) )
        #         except TypeError as msg:
        #             raise TypeError( "malformatted data: expected scalar, got '%s'; msg=%s" % (str(work[row][column]), msg) )
        

# for compatibility
Matrix = TableMatrix

class NumpyMatrix( TableBase, MatrixBase ): 
    """A nxm matrix.

    It implements column-wise and row-wise transformations.

    This class adds the following options to the :term:`report` directive.

        :term:`transform-matrix`: apply a matrix transform. Possible choices are:

           * *correspondence-analysis*: use correspondence analysis to permute rows/columns 
           * *normalized-column-total*: normalize by column total
           * *normalized-column-max*: normalize by column maximum
           * *normalized-row-total*: normalize by row total
           * *normalized-row-max*: normalize by row maximum
           * *normalized-total*: normalize over whole matrix
           * *normalized-max*: normalize over whole matrix
           * *sort* : sort matrix rows and columns alphanumerically.
           * filter-by-rows: only take columns that are also present in rows
           * filter-by-cols: only take columns that are also present in cols
           * square: make square matrix (only take rows and columns present in both)
           * *add-row-total* : add the row total at the bottom
           * *add-column-total* : add the column total as a last column

    Requires one level with three fields:

       rows
       columns
       matrix

    All values need to be numerical.
    """
    
    nlevels = 1

    options = TableBase.options +\
        ( ('transform-matrix', directives.unchanged), )

    def __init__(self, *args, **kwargs):

        TableBase.__init__(self, *args, **kwargs )
        MatrixBase.__init__(self, *args, **kwargs )

    def buildMatrix( self, 
                     work, 
                     missing_value = 0, 
                     apply_transformations = True,
                     take = None,
                     dtype = numpy.float ):
        """build a matrix from work, a single level dictionary with 
        three fields: rows, columns, matrix."""
        
        try: rows = work["rows"]
        except KeyError: raise KeyError( "expected rownames in field 'rows' - no 'rows' present: %s " % \
                                             str(list(work.keys()) ))
        
        try: columns = work["columns"]
        except KeyError: raise KeyError( "expected column names in field 'columns' - no 'columns' present: %s " % \
                                             str(list(work.keys())) )

        try: matrix = work["matrix"]
        except KeyError: raise KeyError( "expected matrix in field 'matrix' - no 'matrix' present: %s" % \
                                             str(list(work.keys())) )
        
        nrows, ncolumns = matrix.shape

        if len(rows) != nrows:
            raise ValueError("number of rows does not correspond to matrix: %i != %i" % (len(rows), nrows))

        if len(columns) != ncolumns:
            raise ValueError("number of columns does not correspond to matrix: %i != %i" % (len(columns), ncolumns))

        # convert rows/columns to str (might be None)
        rows = [ str(x) for x in rows ]
        columns = [ str(x) for x in columns ]

        if self.converters and apply_transformations:
            for converter in self.converters: 
                self.debug("applying converter %s" % converter)
                matrix, rows, columns = converter(matrix, rows, columns)
        
        return matrix, rows, columns

class Debug( Renderer ):
    '''a simple renderer, returning the type of data and the number of items at each path.'''

    # only look at leaves
    nlevels = 1

    def render( self, dataseries, path ):
        
        print dataseries
        print path
        # initiate output structure
        results = ResultBlocks( title = path )

        # iterate over all items at leaf
        results.append( ResultBlock( "path= %s, data= %s" % \
                                         ( path2str(path),
                                           dataseries), 
                                     title = "") )

        return results
        
class User(Renderer):
    """Renderer for user-implemented rendering.

    The renderer itself creates no output, but returns the results
    of the tracker.

    When called, a Renderer and its subclasses will return blocks of
    restructured text. Images are automatically collected from matplotlib
    and other renderers from active graphics devices and inserted at 
    the place-holders.
    """

    # only leaves
    nlevels = 1

    def render( self, work, path ):
        
        # initiate output structure
        results = ResultBlocks( title = path2str(path) )

        # iterate over all items at leaf
        #for key in work:

        for key in Utils.TrackerKeywords:
            if key in work:
                # add a result block
                results.append( ResultBlock( work[key],
                                             title = "" ) )
            
        return results

class Status( Renderer ):
    '''Renders a status report.

    A status report is a two element table containing
    status ('PASS', 'FAIL', 'WARNING', 'NA') and some information.

    The __doc__ string of the tracker is added as a legend below
    the status report.
    '''

    # read complete data
    nlevels = -1

    map_code2image = { 'FAIL' : "fail.png",
                       'PASS' : "pass.png",
                       'NA': "not_available.png",
                       'WARNING' : "warning.png",
                       'WARN': "warning.png" }

    def __call__(self, data, path ):

        lines = []
        dirname = os.path.join( os.path.dirname(sys.modules["SphinxReport"].__file__), "images" )
        descriptions = {}
        title = "status"

        # add header
        lines.append( ".. csv-table:: %s" % "table" )
        lines.append( "   :class: sortable" )
        lines.append( '   :header: "Track", "Test", "", "Status", "Info" ' )
        lines.append( '' )

        for testname, w in data.items():
            for track, work in w.items():
            
                status = str(work['status']).strip()
                descriptions[testname] = work.get('description', "")
                info = str(work['info']).strip()
                try:
                    image = ".. image:: %s" % os.path.join( dirname, self.map_code2image[status.upper()] )
                except KeyError:
                    image = ""

                lines.append( '   "%(track)s",":term:`%(testname)s`","%(image)s","%(status)s","%(info)s"' % locals() )
                
        lines.append( "") 
        
        lines.append( ".. glossary::" )
        lines.append( "" )

        for test, description in descriptions.items():
            lines.append( '%s\n%s\n' % (Utils.indent(test,3), Utils.indent( description,6) ) )
        
        return ResultBlocks( ResultBlock( "\n".join(lines), title = "") )        



