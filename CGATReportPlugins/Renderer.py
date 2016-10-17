import os
import sys
import re
import json
import io
from docutils.parsers.rst import directives
import numpy
import pandas
# for output of excel work books
import openpyxl

# import numpy.arange and math.log into namepsace that they
# are available in eval
from numpy import arange
from math import log

from CGATReport.ResultBlock import ResultBlock, ResultBlocks
from CGATReport.DataTree import path2str
from CGATReport import Utils, DataTree
from CGATReport.Component import Component
from CGATReport import CorrespondenceAnalysis


class Renderer(Component):
    """Base class of renderers that render data into restructured text.

    The subclasses define how to render the data by overloading the
    :meth:`render` method.

    When called, a Renderer and its subclasses will return blocks of
    restructured text. Images are automatically collected from
    matplotlib and inserted at place-holders.

    """
    # plugin fields
    capabilities = ["render"]

    options = (('format', directives.unchanged),
               ('split-at', directives.nonnegative_int),
               ('split-always', directives.unchanged),)

    # required levels in DataTree
    # -1: Renderer can work from any number of levels
    # 0: Renderer needs scalar
    # 1: Renderer requires dataframe with 1 row-label
    # 2: Renderer requires dataframe with 2 row-labels
    # ...
    nlevels = None

    # default number format
    format = "%i"

    # split number of tracks
    split_at = 0

    # tracks always to include if plot is split into
    # several plots.
    split_always = []

    # when splitting, keep first column
    split_keep_first_column = False

    # default group_level for render
    group_level = 0

    # directory where the actual restructured text is located.
    rst_dir = None

    # directory of the root of the source document. rstdir is
    # a subdirectory of this.
    src_dir = None

    # directory in which the document is built.
    build_dir = None

    # dictionary of rst options controlling the display of a
    # directive.
    display_options = {}

    def __init__(self, *args, **kwargs):
        """create an Renderer object using an instance of
        a:class:`Tracker.Tracker`.
        """
        Component.__init__(self, *args, **kwargs)

        self.debug("%s: starting renderer '%s'" %
                   (id(self), str(self)))

        try:
            self.format = kwargs["format"]
        except KeyError:
            pass

        self.split_at = int(kwargs.get("split-at", 0))

        if "split-always" in kwargs:
            self.split_always = kwargs["split-always"].split(',')

    def __call__(self, dataframe, path):
        '''iterate over leaves/branches in data structure.

        This method will call the:meth:`render` method.

        Large dataframes are split into multiple, smaller rendered
        objects if self.split_at is not zero.

        By default, dataframes are split along the hierachical
        index. However, if there is only a single index, but multiple
        columns, the split is performed on the columns instead. This
        is used when splitting coordinate data as a result of the
        histogram transformation.

        '''
        result = ResultBlocks()

        if not self.split_at:
            result.extend(self.render(dataframe, path))
        else:
            # split dataframe at first index
            level = Utils.getGroupLevels(dataframe)
            grouper = dataframe.groupby(level=level)

            # split dataframe column wise if only one index
            # and multiple columns
            if len(grouper) == 1 and len(dataframe.columns) > self.split_at:
                columns = list(dataframe.columns)
                always = []
                if self.split_keep_first_column:
                    always.append(columns[0])
                # columns to always keep
                always.extend([c for c in columns if c in self.split_always])
                columns = [c for c in columns if c not in always]
                for x in range(0, len(columns), self.split_at):
                    # extract a set of columns
                    result.extend(self.render(
                        dataframe.loc[:, always+columns[x:x+self.split_at]],
                        path))
            # split dataframe along index
            elif len(grouper) >= self.split_at:
                # build groups
                always, remove_always = [], set()

                if self.split_always:
                    for key, work in grouper:
                        for pat in self.split_always:
                            rx = re.compile(pat)
                            if rx.search(path2str(key)):
                                always.append((key, work))
                                remove_always.add(key)

                    grouper = dataframe.groupby(level=level)

                def _group_group(grouper, always, remove_always):
                    group = always[:]
                    for key, work in grouper:

                        if key in remove_always:
                            continue
                        group.append((key, work))

                        if len(group) >= self.split_at:
                            yield group
                            group = always[:]

                    # reconcile index names
                    yield group

                first = True
                for group in _group_group(grouper,
                                          always,
                                          remove_always):
                    # do not plot last dataframe that contains
                    # only the common tracks to plot
                    if not first and len(group) == len(always):
                        continue
                    first = False

                    df = pandas.concat(
                        [x[1] for x in group])

                    # reconcile index names
                    df.index.names = dataframe.index.names
                    result.extend(self.render(df, path))
            else:
                # do not split dataframe
                result.extend(self.render(dataframe, path))

        return result

    def toString(self, value):
        '''returns a number as string

        If not a number, return empty string.'''

        try:
            return self.format % value
        except TypeError:
            return ""
        except ValueError:
            return "nan"

    def set_paths(self, rst_dir, src_dir, build_dir):
        '''set document paths. These are relevant to determine the
        relative position of the object to be rendered in a hierarchy
        of documents.
        '''
        self.rst_dir = rst_dir
        self.src_dir = src_dir
        self.build_dir = build_dir

    def get_paths(self):
        return self.rst_dir, self.src_dir, self.build_dir

    def set_display_options(self, display_options):
        '''set display options given by user.'''
        self.display_options = display_options


class DataFrame(Renderer):
    """
    Renderer outputting the dataframe directly
    as text without any modification.
    """

    options = Renderer.options +\
        (('head', directives.length_or_unitless),
         ('tail', directives.length_or_unitless),
         ('summary', directives.unchanged))

    # output head of dataframe - 0 is no head
    head = 0
    # output tail of dataframe - 0 is no tail
    tail = 0
    # output summary of dataframe
    summary = False

    def __init__(self, *args, **kwargs):

        Renderer.__init__(self, *args, **kwargs)
        self.head = int(kwargs.get('head', 0))
        self.tail = int(kwargs.get('tail', 0))
        self.summary = 'summary' in kwargs

    def __call__(self, dataframe, path):
        result = ResultBlocks()
        texts = []
        if self.head or self.tail:
            if self.head:
                texts.append(str(dataframe.head(self.head)))
            if self.tail:
                texts.append(str(dataframe.tail(self.tail)))
        elif self.summary:
            texts.append(str(dataframe.describe()))
        else:
            texts.append(str(dataframe))

        # add indentation
        texts = ['\n'.join(['   %s' % y for y in x.split('\n')])
                 for x in texts]

        formatted = '''
::

%s

''' % '\n   ...\n'.join(texts)

        result.append(ResultBlock(formatted,
                                  title=path2str(path)))
        return result


class TableBase(Renderer):

    '''base classes for tables and matrices.'''

    options = Renderer.options +\
        (('force', directives.unchanged),
         ('separate', directives.unchanged),
         ('max-rows', directives.length_or_unitless),
         ('large-rows', directives.length_or_unitless),
         ('max-cols', directives.length_or_unitless),
         ('large-html-class', directives.unchanged),
         ('add-rowindex', directives.unchanged),
         ('html-class', directives.unchanged))

    max_rows = 50
    max_cols = 20

    def __init__(self, *args, **kwargs):
        Renderer.__init__(self, *args, **kwargs)

        self.force = "force" in kwargs
        self.separate = "separate" in kwargs
        self.max_rows = kwargs.get("max-rows", 50)
        self.max_cols = kwargs.get("max-cols", 20)
        self.add_rowindex = "add-rowindex" in kwargs
        self.html_class = kwargs.get(
            'html-class',
            Utils.PARAMS.get('report_table_class',
                             None))
        self.large_rows = kwargs.get("max-rows", 20)
        self.large_html_class = kwargs.get(
            'large-html-class',
            Utils.PARAMS.get('report_largetable_class',
                             None))

    def asCSV(self, dataframe, row_headers, col_headers, title):
        '''save the table using CSV.'''

        out = io.StringIO()
        dataframe.to_csv(out)
        result = []
        result.append(".. csv-table:: %s" % title)
        lines = out.getvalue().split("\n")

        if len(lines) > self.large_rows:
            if self.large_html_class is not None:
                result.append("   :class: %s" % self.large_html_class)
        else:
            if self.html_class is not None:
                result.append("   :class: %s" % self.html_class)

        if self.add_rowindex:
            raise NotImplementedError('add-rowindex not implemented')
        else:
            result.append('   :header: %s' % lines[0])
            result.append('')
            result.extend(['   %s' % x for x in lines[1:]])

        result.append("")

        return ResultBlock("\n".join(result), title=title)

    def asRST(self, dataframe, row_headers, col_headers, title):
        '''save the table using RST.'''

        out = io.StringIO()
        dataframe.to_csv(out)
        data = [x.split(',') for x in out.getvalue().split('\n')]
        # ignore last element - empty
        del data[-1]

        ncols = len(data[0])
        max_widths = [max([len(row[x]) for row in data])
                      for x in range(ncols)]

        separator = "+" + "+".join(["-" * (x + 2) for x in max_widths]) + "+"
        format = "|" + "|".join([" %%%is " % x for x in max_widths]) + "|"

        lines = []
        lines.append(separator)
        for row in data:
            lines.append(format % tuple(map(str, row)))
            lines.append(separator)
        lines.append("")

        return ResultBlock("\n".join(lines), title=title)

    def asFile(self, dataframe, row_headers, col_headers, title):
        '''save the table as HTML file.

        Multiple files of the same Renderer/Tracker combination are
        distinguished by the title.
        '''

        self.debug("%s: saving %i x %i table as file'" %
                   (id(self),
                    len(row_headers),
                    len(col_headers)))
        lines = []
        lines.append("`%i x %i table <#$html %s$#>`__" %
                     (len(row_headers), len(col_headers),
                      title))

        r = ResultBlock("\n".join(lines) + "\n", title=title)

        out = io.StringIO()
        dataframe.to_csv(out)
        lines = out.getvalue().split("\n")

        # create an html table
        data = ["<table>"]
        data.append("<tr><th></th><th>%s</th></tr>" %
                    "</th><th>".join(map(str, lines[0].split(","))))
        data.extend(["<tr><td>%s</td></tr>" %
                     ("</td><td>".join(x.split(","))) for x in lines[1:]])
        data.append("</table>\n")

        # substitute links
        data = [re.sub(
            "`(.*?(?:\".+\"|\'.+\')?.*?)\s<(.*?(?:\".+\"|\'.+\')?.*?)>`_",
            r'<a href="\2">\1</a>',
            x)
            for x in data]

        r.html = "\n".join(data)

        return r

    def asSpreadSheet(self, dataframe, row_headers, col_headers, title):
        '''save the table as an xls file.

        Multiple files of the same Renderer/Tracker combination are
        distinguished by the title.
        '''

        self.debug("%s: saving %i x %i table as spread-sheet'" %
                   (id(self),
                    len(row_headers),
                    len(col_headers)))

        is_hierarchical = isinstance(dataframe.index,
                                     pandas.core.index.MultiIndex)

        split = is_hierarchical and len(dataframe.index.levels) > 1

        quick = len(dataframe) > 10000
        if quick and not split:
            # quick writing, only append method works
            wb = openpyxl.Workbook(optimized_write=True)

            def fillWorksheet(ws, dataframe, title):
                ws.append([""] + list(col_headers))
                for x, row in enumerate(dataframe.iterrows()):
                    ws.append([path2str(row[0])] + list(row[1]))

                # patch: maximum title length seems to be 31
                ws.title = title[:30]

        else:
            # do it cell-by-cell, this might be slow
            wb = openpyxl.Workbook(optimized_write=False)

            def fillWorksheet(ws, dataframe, title):
                # regex to detect rst hypelinks
                regex_link = re.compile('`(.*) <(.*)>`_')
                # write row names
                for row, row_name in enumerate(dataframe.index):
                    # rows and columns start at 1
                    c = ws.cell(row=row + 2, column=1)
                    c.value = row_name

                # write columns
                for column, column_name in enumerate(dataframe.columns):
                    # set column title
                    # rows and columns start at 1
                    c = ws.cell(row=1, column=column + 2)
                    c.value = column_name

                    # set column values
                    dataseries = dataframe[column_name]

                    if dataseries.dtype == object:
                        for row, value in enumerate(dataseries):
                            c = ws.cell(row=row + 2,
                                        column=column + 2)
                            value = str(value)
                            if value.startswith('`'):
                                c.value, c.hyperlink =\
                                    regex_link.match(value).groups()
                            else:
                                c.value = value
                    else:
                        for row, value in enumerate(dataseries):
                            c = ws.cell(row=row + 2,
                                        column=column + 2)
                            c.value = value
                if title:
                    # patch: maximum title length seems to be 31
                    ws.title = re.sub("/", "_", title)[:30]

        if len(wb.worksheets) == 0:
            wb.create_sheet()

        if split:
            # create separate worksheets for nested indices
            nlevels = len(dataframe.index.levels)
            paths = list(map(tuple, DataTree.unique(
                [x[:nlevels - 1]
                 for x in dataframe.index.unique()])))

            ws = wb.worksheets[0]
            ws.title = 'Summary'
            ws.append(
                [""] * (nlevels - 1) + ["Worksheet", "Rows"])

            for row, path in enumerate(paths):
                # select data frame as cross-section
                work = dataframe.xs(path, axis=0)
                title = path2str(path)
                if len(title) > 30:
                    title = "sheet%i" % row

                ws.append(list(path) + [title, len(work)])
                c = ws.cell(row=row + 1,
                            column=nlevels)
                # this does not work in oocalc
                c.hyperlink = "#%s!A1" % title
                fillWorksheet(wb.create_sheet(),
                              work,
                              title=title)
        else:
            fillWorksheet(wb.worksheets[0], dataframe,
                          title=title)

        # write result block
        lines = []
        lines.append("`%i x %i table <#$xls %s$#>`__" %
                     (len(row_headers), len(col_headers),
                      title))
        lines.append("")

        r = ResultBlock("\n".join(lines), title=title)
        r.xls = wb

        self.debug("%s: saved %i x %i table as spread-sheet'" %
                   (id(self),
                    len(row_headers),
                    len(col_headers)))
        return r


class Table(TableBase):

    '''a basic table.

    Values are either text or converted to text.

    If the has more rows than:attr:`max_rows` or more columns
    than:attr:`max_cols`, a placeholder will be inserted pointing
    towards a file.

    The attribute:attr:`large` determines where large tables are
    written to. The default is html. Alternative values are ``xls``
    for excel spread-sheets.

    '''
    options = TableBase.options +\
        (('transpose', directives.unchanged),
         ('add-percent', directives.unchanged),
         ('head', directives.length_or_unitless),
         ('large', directives.unchanged),
         ('format-columns', directives.unchanged),
         ('preview', directives.unchanged))

    # By default, group everything together
    group_level = "all"

    transpose = False

    def __init__(self, *args, **kwargs):
        TableBase.__init__(self, *args, **kwargs)

        self.transpose = "transpose" in kwargs
        self.large = kwargs.get("large", "html")
        self.preview = "preview" in kwargs

        if "format-columns" in kwargs:
            self.format_columns = kwargs.get("format-columns",
                                             None)
            if not self.format_columns:
                self.format_columns = "auto"
            else:
                columns = self.format_columns.split(",")
                self.format_columns = {}
                for column in columns:
                    if "=" in column:
                        column, fmt = column.split("=")
                    else:
                        fmt = "auto"
                    self.format_columns[column] = fmt
        else:
            self.format_columns = None
        
        self.add_percent = kwargs.get('add-percent', None)
        self.head = int(kwargs.get('head', 0))

    def modifyTable(self, dataframe):
        '''modify table if required, for example adding percentages.
        '''

        if self.head > 0:
            dataframe = dataframe[:self.head]

        if self.add_percent:
            columns = dataframe.columns
            parts = self.add_percent.split(";")
            for part in parts:
                total, other_col = None, None
                if "," in part:
                    column, method = part.split(",")
                    if method in columns:
                        other_col = columns.index(method)
                    else:
                        try:
                            total = float(method)
                        except ValueError:
                            raise ValueError(
                                ("can't compute total from "
                                 "expression `%s` - missing column?") %
                                method)
                else:
                    column = part

                if column not in columns:
                    raise ValueError("unknown column `%s` to add-percent" %
                                     (column))

                values = numpy.array(dataframe.sum(axis=1),
                                     dtype=numpy.float)

                if other_col:
                    dataframe['%s/%%' % column] = 100.0 *\
                        dataframe[column] / dataframe[other_col]
                else:
                    if total is None:
                        total = float(sum(values))
                    dataframe['%s/%%' % column] = 100.0 *\
                        dataframe[column] / values

        if self.transpose:
            dataframe = dataframe.transpose()
            # flatten the column index if it is hierarchical
            is_hierarchical = isinstance(dataframe.columns,
                                         pandas.core.index.MultiIndex)
            if is_hierarchical:
                dataframe.columns = ["/".join(x) for x in dataframe.columns]

        if self.format_columns is not None:
            for column in dataframe.columns:
                if self.format_columns == "auto":
                    fmt = "auto"
                elif column in self.format_columns:
                    fmt = self.format_columns[column]
                else:
                    continue

                cc = dataframe[column]
                if cc.dtype == numpy.float:
                    mi, ma = cc.min(), cc.max()
                    if mi >= 0 and ma <= 1.0:
                        # convert to percent
                        dataframe[column] = (100.0 * cc).map('{:,.2f}'.format)
                    else:
                        dataframe[column] = cc.map('{:,.2f}'.format)
                elif cc.dtype == numpy.int:
                    dataframe[column] = cc.map('{:,}'.format)
                    
        # if index is not hierarchical, but contains tuples,
        # split tuples in index to build a new (hierarchical) index
        is_hierarchical = isinstance(dataframe.index,
                                     pandas.core.index.MultiIndex)

        if not is_hierarchical and isinstance(dataframe.index[0], tuple):
            idx = pandas.MultiIndex.from_tuples(dataframe.index)
            dataframe.index = idx

        return dataframe

    def __call__(self, dataframe, path):

        # modify table (adding/removing columns) according to user options
        # matrix, row_headers, col_headers = \
        # self.modifyTable(matrix, row_headers, col_headers)
        dataframe = self.modifyTable(dataframe)

        title = path2str(path)

        results = ResultBlocks()

        row_headers = dataframe.index
        col_headers = dataframe.columns

        # as of sphinx 1.3.1, tables with more than 100 columns cause an
        # error:
        # Exception occurred:
        # File "/ifs/apps/apps/python-2.7.9/lib/python2.7/site-packages/docutils/writers/html4css1/__init__.py", line 642, in write_colspecs
        # colwidth = int(node['colwidth'] * 100.0 / width + 0.5)
        # ZeroDivisionError: float division by zero
        #
        # Thus, for table with more than 100 columns, force will be
        # disabled and max_cols set to a low value in order to make
        # sure the table is not displayed inline
        if len(col_headers) >= 90:
            self.force = False
            self.max_cols = 10

        # do not output large matrices as rst files
        if self.separate or (not self.force and
                             (len(row_headers) > self.max_rows or
                              len(col_headers) > self.max_cols)):
            if self.large == "xls":
                results.append(self.asSpreadSheet(dataframe, row_headers,
                                                  col_headers, title))
            else:
                results.append(self.asFile(dataframe, row_headers,
                                           col_headers, title))

            if self.preview:
                raise NotImplementedError('preview not implemented')
                row_headers = row_headers[:self.max_rows]
                col_headers = col_headers[:self.max_cols]
                # matrix = [x[:self.max_cols] for x in
                #          matrix[:self.max_rows]]
            else:
                return results

        results.append(self.asCSV(dataframe, row_headers, col_headers, title))

        return results


class HTMLTable(Table):

    '''a table in html format to download.
    '''

    def __call__(self, dataframe, path):

        results = ResultBlocks()
        if dataframe is None:
            return results

        title = path2str(path)
        row_headers = dataframe.index
        col_headers = dataframe.columns
        results = ResultBlocks()
        results.append(self.asFile(dataframe, row_headers,
                                   col_headers, title))

        return results


class XlsTable(Table):

    '''a table in xls format to download.
    '''

    def __call__(self, dataframe, path):

        results = ResultBlocks()
        if dataframe is None:
            return results

        title = path2str(path)
        row_headers = dataframe.index
        col_headers = dataframe.columns
        results.append(self.asSpreadSheet(dataframe, row_headers,
                                          col_headers, title))
        
        return results


class RstTable(Table):

    '''an rst formatted table.
    '''

    def __call__(self, dataframe, path):

        results = ResultBlocks()

        if dataframe is None:
            return results

        title = path2str(path)

        row_headers = dataframe.index
        col_headers = dataframe.columns

        # do not output large matrices as rst files
        if self.separate or (not self.force and
                             (len(row_headers) > self.max_rows or
                              len(col_headers) > self.max_cols)):
            if self.large == "xls":
                results.append(self.asSpreadSheet(dataframe, row_headers,
                                                  col_headers, title))
            else:
                results.append(self.asFile(dataframe, row_headers,
                                           col_headers, title))

        results.append(self.asRST(dataframe, row_headers, col_headers, title))

        return results


class GlossaryTable(Table):

    """output a table in the form of a glossary."""

    def __call__(self, dataframe, path):

        results = ResultBlocks()

        if dataframe is None:
            return results

        title = path2str(path)

        lines = []
        lines.append(".. glossary::")
        lines.append("")

        for x, row in enumerate(dataframe.iterrows()):
            header, data = row
            txt = "\n      ".join([x.strip() for x in str(data).split("\n")])
            lines.append('   %s\n      %s\n' % (path2str(header), txt))

        lines.append("")

        results.append(ResultBlock("\n".join(lines), title=title))

        return results


class MatrixBase:
    '''base class for matrices.

    This base class provides utility functions for rectangular 2D matrices.

    It implements column-wise and row-wise transformations.

    This class adds the following options to the:term:`report` directive.

    :term:`transform-matrix`: apply a matrix transform. Possible choices are:

           * *correspondence-analysis*: use correspondence analysis
             to permute rows/columns
           * *normalized-col-total*: normalize by column total
           * *normalized-col-max*: normalize by column maximum,
           * *normalized-col-first*: normalize by first column.
             The first column is then removed.
           * *normalized-row-total*: normalize by row total
           * *normalized-row-max*: normalize by row maximum
           * *normalized-row-first*: normalize by first row.
             The first row is then removed.
           * *normalized-total*: normalize over whole matrix
           * *normalized-max*: normalize over whole matrix
           * *sort*: sort matrix rows and columns alphanumerically.
           * *sort-numerically*: sort matrix rows and columns numerically.
           * filter-by-rows: only take columns that are also present in rows
           * filter-by-cols: only take columns that are also present in cols
           * square: make square matrix (only take rows and columns present
             in both)
           * *add-row-total*: add the row total at the bottom
           * *add-column-total*: add the column total as a last column

    Requires two levels:

       rows[dict] / columns[dict] / value

    All values need to be numerical.
    '''

    def __init__(self, *args, **kwargs):

        self.mMapKeywordToTransform = {
            "correspondence-analysis": self.transformCorrespondenceAnalysis,
            "transpose": self.transformTranspose,
            "normalized-row-total": self.transformNormalizeRowTotal,
            "normalized-row-max": self.transformNormalizeRowMax,
            "normalized-row-first": self.transformNormalizeRowFirst,
            "normalized-col-total": self.transformNormalizeColumnTotal,
            "normalized-col-max": self.transformNormalizeColumnMax,
            "normalized-col-first": self.transformNormalizeColumnFirst,
            "normalized-total": self.transformNormalizeTotal,
            "normalized-max": self.transformNormalizeMax,
            "symmetric-max": self.transformSymmetricMax,
            "symmetric-min": self.transformSymmetricMin,
            "symmetric-avg": self.transformSymmetricAverage,
            "symmetric-sum": self.transformSymmetricSum,
            "filter-by-rows": self.transformFilterByRows,
            "filter-by-cols": self.transformFilterByColumns,
            "square": self.transformSquare,
            "add-row-total": self.transformAddRowTotal,
            "add-column-total": self.transformAddColumnTotal,
            "sort": self.transformSort,
            "sort-numerically": self.transformSortNumerically,
        }
        self.tofloat = False
        self.converters = []
        if "transform-matrix" in kwargs:
            for kw in [x.strip() for x in
                       kwargs["transform-matrix"].split(",")]:
                if kw.startswith("normalized"):
                    self.format = "%6.4f"
                    self.tofloat = True
                try:
                    self.converters.append(self.mMapKeywordToTransform[kw])
                except KeyError:
                    raise ValueError(("unknown matrix transformation %s, "
                                      "possible values are: %s") %
                                     (kw,
                                      ",".join(sorted(
                                          self.mMapKeywordToTransform.keys()
                                      ))))

    def transformAddRowTotal(self, matrix, row_headers, col_headers):
        '''add row total to the matrix.'''
        nrows, ncols = matrix.shape

        totals = numpy.zeros(nrows)
        for x in range(nrows):
            totals[x] = sum(matrix[x, :])
        col_headers.append("total")
        totals.shape = (nrows, 1)
        return (numpy.hstack(numpy.hsplit(matrix, ncols) + [totals, ]),
                row_headers, col_headers)

    def transformAddColumnTotal(self, matrix, row_headers, col_headers):
        raise NotImplementedError

    def transformFilterByRows(self, matrix, row_headers, col_headers):
        """only take columns that are also present in rows"""
        take = [x for x in range(len(col_headers))
                if col_headers[x] in row_headers]
        return (matrix.take(take, axis=1), row_headers,
                [col_headers[x] for x in take])

    def transformFilterByColumns(self, matrix, row_headers, col_headers):
        """only take rows that are also present in columns"""
        take = [x for x in range(len(row_headers))
                if row_headers[x] in col_headers]
        return (matrix.take(take, axis=0),
                [row_headers[x] for x in take],
                col_headers)

    def transformSquare(self, matrix, row_headers, col_headers):
        """only take rows and columns that are present in both giving a square
        matrix.
        """
        take = set(row_headers).intersection(set(col_headers))
        row_indices = [x for x in range(len(row_headers))
                       if row_headers[x] in take]
        col_indices = [x for x in range(len(col_headers))
                       if col_headers[x] in take]
        m1 = matrix.take(row_indices, axis=0)
        m2 = m1.take(col_indices, axis=1)

        return (m2,
                [row_headers[x] for x in row_indices],
                [col_headers[x] for x in col_indices])

    def transformTranspose(self, matrix, row_headers, col_headers):
        """transpose the matrix."""
        return numpy.transpose(matrix), col_headers, row_headers

    def transformCorrespondenceAnalysis(self, matrix,
                                        row_headers,
                                        col_headers):
        """apply correspondence analysis to a matrix.
        """
        if len(row_headers) <= 1 or len(col_headers) <= 1:
            self.warn("correspondence analysis skipped for "
                      "matrices with single row/column")
            return matrix, row_headers, col_headers

        try:
            row_indices, col_indices = \
                CorrespondenceAnalysis.GetIndices(matrix)
        except ValueError:
            return matrix, row_headers, col_headers

        map_row_new2old = numpy.argsort(row_indices)
        map_col_new2old = numpy.argsort(col_indices)

        matrix, row_headers, col_headers = \
            CorrespondenceAnalysis.GetPermutatedMatrix(matrix,
                                                       map_row_new2old,
                                                       map_col_new2old,
                                                       row_headers=row_headers,
                                                       col_headers=col_headers)

        return matrix, row_headers, col_headers

    def transformSort(self, matrix, row_headers, col_headers):
        """apply correspondence analysis to a matrix.
        """

        map_row_new2old = [x[0] for x in sorted(enumerate(row_headers),
                                                key=lambda x: x[1])]
        map_col_new2old = [x[0] for x in sorted(enumerate(col_headers),
                                                key=lambda x: x[1])]

        matrix, row_headers, col_headers = \
            CorrespondenceAnalysis.GetPermutatedMatrix(matrix,
                                                       map_row_new2old,
                                                       map_col_new2old,
                                                       row_headers=row_headers,
                                                       col_headers=col_headers)

        return matrix, row_headers, col_headers

    def transformSortNumerically(self, matrix, row_headers, col_headers):
        """apply correspondence analysis to a matrix.
        """

        def atoi(text):
            return int(text) if text.isdigit() else text

        def natural_keys(text):
            '''
            alist.sort(key=natural_keys) sorts in human order
            http://nedbatchelder.com/blog/200712/human_sorting.html
            (See Toothy's implementation in the comments)
            '''
            return [atoi(c) for c in re.split('(\d+)', text)]
            
        map_row_new2old = [x[0] for x in sorted(
            enumerate([natural_keys(x) for x in row_headers]),
            key=lambda x: x[1])]

        map_col_new2old = [x[0] for x in sorted(
            enumerate([natural_keys(x) for x in col_headers]),
            key=lambda x: x[1])]

        matrix, row_headers, col_headers = \
            CorrespondenceAnalysis.GetPermutatedMatrix(matrix,
                                                       map_row_new2old,
                                                       map_col_new2old,
                                                       row_headers=row_headers,
                                                       col_headers=col_headers)

        return matrix, row_headers, col_headers            

    def transformSymmetricMax(self, matrix, rows, cols):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError("matrix not square - can not be symmetrized")

        for x in range(len(rows)):
            for y in range(x + 1, len(cols)):
                matrix[x, y] = matrix[y, x] = max(matrix[x, y], matrix[y, x])

        return matrix, rows, cols

    def transformSymmetricMin(self, matrix, rows, cols):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError("matrix not square - can not be symmetrized")

        for x in range(len(rows)):
            for y in range(x + 1, len(cols)):
                matrix[x, y] = matrix[y, x] = \
                    min(matrix[x, y], matrix[y, x])

        return matrix, rows, cols

    def transformSymmetricSum(self, matrix, rows, cols):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError("matrix not square - can not be symmetrized")

        for x in range(len(rows)):
            for y in range(x + 1, len(cols)):
                matrix[x, y] = matrix[y, x] = \
                    sum(matrix[x, y], matrix[y, x])

        return matrix, rows, cols

    def transformSymmetricAverage(self, matrix, rows, cols):
        """symmetrize a matrix.

        returns the normalized matrix.
        """
        if len(rows) != len(cols):
            raise ValueError("matrix not square - can not be symmetrized")

        for x in range(len(rows)):
            for y in range(x + 1, len(cols)):
                matrix[x, y] = matrix[y, x] = \
                    sum(matrix[x, y], matrix[y, x]) / 2

        return matrix, rows, cols

    def transformNormalizeTotal(self, matrix, rows, cols):
        """normalize a matrix by the total.

        Returns the normalized matrix.
        """
        return matrix / matrix.sum(), rows, cols

    def transformNormalizeMax(self, matrix, rows, cols):
        """normalize a matrix by the max.

        Returns the normalized matrix.
        """
        return matrix / matrix.max(), rows, cols

    def transformNormalizeRowTotal(self, matrix, rows, cols):
        """normalize a matrix row by the row total.

        Returns the normalized matrix.
        """
        return matrix / matrix.sum(axis=1), rows, cols

    def transformNormalizeRowMax(self, matrix, rows, cols):
        """normalize a matrix row by the row maximum.

        Returns the normalized matrix.
        """
        return matrix / matrix.max(axis=1), rows, cols

    def transformNormalizeRowFirst(self, matrix, rows, cols):
        """normalize a matrix row by the row maximum.

        Returns the normalized matrix.
        """
        m = matrix[0]
        matrix = numpy.delete(matrix, 0, 0)
        return matrix / m, rows[1:], cols

    def transformNormalizeColumnTotal(self, matrix, rows, cols):
        """normalize a matrix by the column total.

        Returns the normalized matrix."""
        return matrix / matrix.sum(axis=0), rows, cols

    def transformNormalizeColumnFirst(self, matrix, rows, cols):
        """normalize a matrix by the first column.

        Removes the first column.

        Returns the normalized matrix."""
        m = matrix[:, 0]
        matrix = numpy.delete(matrix, 0, 1)
        return matrix / m, rows, cols[1:]

    def transformNormalizeColumnMax(self, matrix, rows, cols):
        """normalize a matrix by the column maximum

        Returns the normalized matrix.
        """
        return matrix / matrix.max(axis=0), rows, cols

    def render(self, work, path):
        """render the data.
        """

        results = ResultBlocks(title=path)

        matrix, rows, columns = self.buildMatrix(work)

        title = path2str(path)

        if len(rows) == 0:
            return ResultBlocks(ResultBlock("", title=title))

        # do not output large matrices as rst files
        # separate and force need to be mixed in.
        if self.separate or (not self.force and
                             (len(rows) > self.max_rows or
                              len(columns) > self.max_cols)):
            return ResultBlocks(self.asFile(pandas.DataFrame(matrix,
                                                             index=rows,
                                                             columns=columns),
                                            rows,
                                            columns,
                                            title),
                                title=path)

        lines = []
        lines.append(".. csv-table:: %s" % title)
        lines.append('   :header: "track","%s" ' % '","'.join(columns))
        lines.append('')
        for row in range(len(rows)):
            lines.append(
                '   "%s","%s"' %
                (rows[row], '","'.join(
                    [self.toString(x) for x in matrix[row]])))
        lines.append("")

        if path is None:
            subtitle = ""
        else:
            subtitle = path2str(path)

        results.append(ResultBlock("\n".join(lines), title=subtitle))

        return results


class TableMatrix(TableBase, MatrixBase):

    """A table with numerical columns.

       rows[dict] / columns[dict] / value

    All values need to be numerical.
    """

    # By default, group everything together
    group_level = "all"

    options = TableBase.options +\
              (('full-row-labels', directives.unchanged),
               ('transform-matrix', directives.unchanged),)

    # wether or not to convert matrix to float
    tofloat = False

    def __init__(self, *args, **kwargs):

        TableBase.__init__(self, *args, **kwargs)
        MatrixBase.__init__(self, *args, **kwargs)

        self.normalize_row_labels = "full-row-labels" not in kwargs

    def buildMatrix(self,
                    dataframe,
                    missing_value=0,
                    apply_transformations=True,
                    dtype=numpy.float):
        """build a matrix from work, a two-level nested dictionary.

        This method will also apply conversions if apply_transformations
        is set.
        """
        
        if self.normalize_row_labels:
            drop = [x for x, y in enumerate(dataframe.index.labels)
                    if len(set(y)) == 1]
            rows = list(map(path2str, dataframe.index.droplevel(drop)))
        else:
            rows = list(map(path2str, dataframe.index))

        columns = list(dataframe.columns)
        # use numpy.matrix - permits easier broadcasting
        # for normalization.
        matrix = numpy.matrix(dataframe.as_matrix())

        # remove columns with only NaNs. This can happend
        # during the dataframe merging process if the
        # columns are unique to each matrix.
        take = numpy.array(numpy.all(numpy.isnan(matrix), axis=0).flat)
        matrix = matrix[:, ~take]
        columns = list(numpy.array(columns)[~take])

        if self.converters and apply_transformations:
            # convert to float for conversions
            if self.tofloat:
                matrix = numpy.matrix(matrix, dtype=numpy.float)

            for converter in self.converters:
                self.debug("applying converter %s" % converter)
                matrix, rows, columns = converter(matrix, rows, columns)

        # convert rows/columns to str (might be None)
        rows = [str(x) for x in rows]
        columns = [str(x) for x in columns]

        # convert numpy.matrix to numpy.array
        return numpy.asarray(matrix), rows, columns

# for compatibility
Matrix = TableMatrix


class NumpyMatrix(TableMatrix, MatrixBase):

    """Deprecated - not needed any more as equivalent to TableMatrix
    """

    def __init__(self, *args, **kwargs):
        TableMatrix.__init__(self, *args, **kwargs)
        MatrixBase.__init__(self, *args, **kwargs)


class Status(Renderer):
    '''Renders a status report.

    A status report is a two element table containing
    status ('PASS', 'FAIL', 'WARNING', 'NA') and some information.

    The __doc__ string of the tracker is added as a legend below
    the status report.
    '''

    options = Renderer.options +\
        (('columns', directives.unchanged),
         ('no-legend', directives.unchanged),)

    # read complete data
    nlevels = -1

    # For backwards compatibility, group by slice (test function)
    # and not track (data track) by default.
    group_level = 'slice'

    map_code2image = {'FAIL': "fail.png",
                      'PASS': "pass.png",
                      'NA': "not_available.png",
                      'WARNING': "warning.png",
                      'WARN': "warning.png"}

    display_legend = True

    columns = ["track", "test", "image", "status", "info"]

    def __init__(self, *args, **kwargs):

        Renderer.__init__(self, *args, **kwargs)
        if "no-legend" in kwargs:
            self.display_legend = False
        if "columns" in kwargs:
            self.columns = [x.strip() for x in kwargs["columns"].split(",")]

    def __call__(self, dataframe, path):

        # convert to dataframe
        # index has test names
        # columns are description, info, status
        columns = ('description', 'info', 'status', 'name')
        if set(dataframe.columns) != set(columns):
            raise ValueError("invalid columns: expected '%s', got '%s' " %
                             (columns, dataframe.columns))

        lines = []
        dirname = os.path.join(os.path.dirname(
            sys.modules["CGATReport"].__file__), "images")
        descriptions = {}
        title = "status"

        # add header
        lines.append(".. csv-table:: %s" % "table")
        lines.append('   :header: "Track", "Test", "", "Status", "Info"')
        lines.append('')
        rows = []
        for index, values in dataframe.iterrows():
            testname = values['name']
            status = values['status']
            try:
                image = ".. image:: {}\n    :width: 32".format(
                    os.path.join(dirname,
                                 self.map_code2image[status.upper()]))
            except KeyError:
                image = ""

            rows.append({
                "test": testname,
                "description": values["description"],
                "info": values['info'],
                "status": status,
                "track": path2str(index),
                "image": image,
            })
            descriptions[testname] = values["description"]

        # filter and sort table
        table = [self.columns]
        table.extend([[row[x] for x in self.columns] for row in rows])

        lines = Utils.table2rst(table).split("\n")

        if self.display_legend:
            lines.append(".. glossary::")
            lines.append("")

            for test, description in list(descriptions.items()):
                lines.append('%s\n%s\n' % (Utils.indent(test, 3),
                                           Utils.indent(description, 6)))

        return ResultBlocks(ResultBlock("\n".join(lines), title=""))


class StatusMatrix(Status, TableBase):
    '''Renders a status matrix.

    A status matrix is a rectangular matrix displaying icons for
    status flags ('PASS', 'FAIL', 'WARNING', 'NA').

    '''
    options = Status.options + TableBase.options +\
              (('row-column', directives.unchanged),
               ('transpose', directives.unchanged),)

    # For backwards compatibility, group by slice (test function)
    # and not track (data track) by default.
    group_level = None

    transpose = False

    row_column = "track"

    def __init__(self, *args, **kwargs):
        TableBase.__init__(self, *args, **kwargs)
        Status.__init__(self, *args, **kwargs)

        self.transpose = "transpose" in kwargs
        self.row_column = kwargs.get("row-column", "track")

    def __call__(self, dataframe, path):

        dirname = os.path.join(os.path.dirname(
            sys.modules["CGATReport"].__file__), "images")

        dataframe["status"] = [
            ".. image:: {}\n    :width: 32".format(
                os.path.join(dirname,
                             self.map_code2image.get(x.upper(), "NA")))
            for x in dataframe["status"]]

        table = pandas.pivot_table(
            dataframe.reset_index(),
            index=[self.row_column],
            columns="slice",
            values="status",
            aggfunc=lambda x: str(x))

        if self.transpose:
            table = table.transpose()
            # flatten the column index if it is hierarchical
            is_hierarchical = isinstance(table.columns,
                                         pandas.core.index.MultiIndex)
            if is_hierarchical:
                table.columns = ["/".join(x) for x in table.columns]

        results = ResultBlocks()
        results.append(self.asCSV(table, [], [], path2str(path)))
        return results


class DataTreeRenderer(object):
    """Base class for renderers processing a data tree
    directly.

    Note that these are special case renderers.
    """
    # plugin fields
    capabilities = ["render"]

    def __init__(self, *args, **kwargs):
        pass

    def __call__(self, data):
        return self.render(data)


class User(DataTreeRenderer):

    """Renderer for user-implemented rendering.

    The renderer itself creates no output, but returns the results of
    the tracker.

    When called, a Renderer and its subclasses will return blocks of
    restructured text. Images are automatically collected from
    matplotlib and other renderers from active graphics devices and
    inserted at the place-holders.

    """

    def __init__(self, *args, **kwargs):
        DataTreeRenderer.__init__(self, *args, **kwargs)

    def render(self, data):

        # initiate output structure
        results = ResultBlocks(title='user')

        labels = DataTree.getPaths(data)
        # iterate over all items at leaf
        for path, branch in DataTree.getNodes(data, len(labels) - 2):
            for key in Utils.TrackerKeywords:
                if key in branch:
                    # add a result block
                    results.append(ResultBlock(branch[key],
                                               title=path2str(path)))

        return results


class Debug(DataTreeRenderer):

    '''a simple renderer, returning the type of data
    and the number of items at each path.'''

    def render(self, data):

        # initiate output structure
        results = ResultBlocks(title='debug')

        try:
            results.append(ResultBlock(json.dumps(
                data, indent=4), title=''))
        except TypeError:
            results.append(ResultBlock(str(data), title=''))

        return results
